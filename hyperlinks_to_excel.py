import os

import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re

# ------------------------------------------
# Config & Setup
# ------------------------------------------

# Target YorkU Health Page
base_url = "https://www.yorku.ca/health/experiential-education-2/"

# Define fills based on visual system
domain_styles = [
    ("https://globalhealth.yorku.ca/", PatternFill(start_color="DDEEDD", end_color="DDEEDD", fill_type="solid")),  # Green: Global Health
    ("https://shpm.info.yorku.ca/", PatternFill(start_color="F2D5ED", end_color="F2D5ED", fill_type="solid")),     # Pink: SHPM
    ("https://kine.info.yorku.ca/", PatternFill(start_color="D1EEFC", end_color="D1EEFC", fill_type="solid")),     # Blue: Kinesiology
    ("https://nursing.info.yorku.ca/", PatternFill(start_color="FCEADB", end_color="FCEADB", fill_type="solid")),  # Peach: Nursing
    ("https://www.yorku.ca/health/psychology/", PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"))  # Yellow: Psychology
]

external_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")  # Grey: External site
pdf_form_fill = PatternFill(start_color="F9C8A6", end_color="F9C8A6", fill_type="solid")  # Orange: PDFs / Forms
error_font = Font(color="FF0000")  # Red: 404 error
default_fill = None
default_font = Font(color="000000")  # Black

exclude_domains = [
    "facebook.com", "twitter.com", "instagram.com",
    "linkedin.com", "youtube.com", "tiktok.com", "pinterest.com"
]

# ------------------------------------------
# Fetch HTML and prepare workbook
# ------------------------------------------

response = requests.get(base_url)
soup = BeautifulSoup(response.text, 'html.parser')

# Get page title for output file
page_title = soup.title.string.strip() if soup.title else "untitled"
page_title_clean = re.sub(r'[\\/*?:"<>|]', "_", page_title)  # Clean filename

# Generate unique numbered filename
counter = 1
while True:
    filename = f"{counter:03d} - links under {page_title_clean}.xlsx"
    if not os.path.exists(filename):
        break
    counter += 1

# Focus on main content area
main = soup.find('main', {'id': 'content', 'class': 'container'})
if not main:
    print("❌ Main content area not found.")
    exit()

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Categorized Links"
ws.append(["Link Text (Clickable)"])  # Header
ws["A1"].font = Font(bold=True)
ws.column_dimensions[get_column_letter(1)].width = 70

# ------------------------------------------
# Process and write links
# ------------------------------------------

seen_urls = set()
count = 0

for a in main.find_all("a", href=True):
    text = a.get_text(strip=True)
    href = a['href']

    # Skip if no text or href
    if not text or not href:
        continue

    # Skip based on text content (case-insensitive)
    if any(keyword in text.lower() for keyword in ["flickr", "home", "learn more"]):
        continue

    # Resolve full URL and skip if already seen or social
    full_url = urljoin(base_url, href)
    if full_url in seen_urls:
        continue
    if any(domain in full_url for domain in exclude_domains):
        continue

    seen_urls.add(full_url)

    # Default styling
    fill = default_fill
    font = default_font

    # Category: PDF or form
    if full_url.lower().endswith((".pdf", ".doc", ".docx")) or "/forms/" in full_url.lower():
        fill = pdf_form_fill
    else:
        matched = False
        for prefix, style in domain_styles:
            if full_url.startswith(prefix):
                fill = style
                matched = True
                break

        if not matched and "yorku.ca/health/" not in full_url:
            fill = external_fill

    # Check if 404 and mark red text (but keep link)
    try:
        head_resp = requests.head(full_url, allow_redirects=True, timeout=5)
        if head_resp.status_code == 404:
            font = error_font
    except:
        pass  # On failure, skip marking red

    # Write to Excel
    cell = ws.cell(row=ws.max_row + 1, column=1, value=text)
    cell.hyperlink = full_url
    cell.style = "Hyperlink"
    if fill:
        cell.fill = fill
    cell.font = font
    count += 1


# ------------------------------------------
# Save file
# ------------------------------------------

wb.save(filename)
print(f"✅ {count} links saved to '{filename}'")

# ------------------------------------------
# 🎨 Legend Reference (for developers)
# ------------------------------------------
# DDEEDD = light green      → Global Health
# F2D5ED = light pink       → SHPM
# D1EEFC = light blue       → Kinesiology
# FCEADB = light peach      → Nursing
# F9C8A6 = orange           → PDFs / Forms
# FFFF00 = yellow           → Psychology
# DDDDDD = light grey       → External site (not under yorku.ca/health)
# FF0000 = red text color   → 404 / error link
# None     = white/default  → Default YorkU Health pages (we control)
